"""
╔══════════════════════════════════════════════════════════════════════════════════╗
║   ILI PIPELINE ALIGNMENT SYSTEM  v10.0                                         ║
║   4-Layer Sequential Hierarchical  ×  Production Excel Engine                  ║
╠══════════════════════════════════════════════════════════════════════════════════╣
║                                                                                  ║
║  ARSITEKTUR:                                                                     ║
║  ┌─────────────────────────────────────────────────────────────────────────┐    ║
║  │  LAYER 0  Validasi Data                                                  │    ║
║  │  LAYER 1  Valve-to-Valve Odometer Correction                             │    ║
║  │  LAYER 2  Weld-to-Weld Sequential Matching (Two-Pointer)                 │    ║
║  │  LAYER 3  Anomali Matching per Spool (Hungarian per Spool)               │    ║
║  │  LAYER 4  Growth Validation per Matched Pair                             │    ║
║  └─────────────────────────────────────────────────────────────────────────┘    ║
╚══════════════════════════════════════════════════════════════════════════════════╝
"""

from __future__ import annotations

import bisect
import copy
import math
import os
import re
import sys
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy.optimize import linear_sum_assignment

warnings.filterwarnings("ignore")


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 1 — KONSTANTA & THRESHOLD  (tidak ada magic number)
# ══════════════════════════════════════════════════════════════════════════════════

INF: float = 1e9

JOINT_LEN_CONSISTENCY_PCT: float = 0.05

CLOCK_HARD_GATE_HR: float = 1.5
W_ODO: float = 0.5
W_CLK: float = 0.3
W_DEP: float = 0.2

C4_DEPTH_TOLERANCE_PCT_WT: float  = 2.0
C5_LENGTH_DECREASE_RATIO: float   = 0.9
C6_MAX_GROWTH_RATE_PER_YR: float  = 1.5

DEFAULT_MAOP_BAR: float = 70.0
DEFAULT_SMYS_MPA: float = 359.0
DEFAULT_OD_MM:    float = 219.1
DEFAULT_WT_MM:    float = 6.4

GROWTH_DEPTH_SUSPECT_DECREASE:  float = 30.0
GROWTH_DEPTH_IMPOSSIBLE:        float = 50.0
CLOCK_SHIFT_NEIGHBOR_WINDOW:    int   = 5
CLOCK_SHIFT_OUTLIER_THRESHOLD:  float = 45.0

UNIVERSAL_SCHEMA = [
    ("feature_id",             "str",   "ID fitur",            False),
    ("log_distance_m",         "float", "Odometer absolut (m)", True),
    ("joint_number",           "int",   "Nomor joint",         False),
    ("dist_to_us_gw_m",        "float", "Jarak ke US GW (m)",  False),
    ("joint_length_m",         "float", "Panjang joint (m)",   False),
    ("feature_type",           "str",   "Tipe fitur",          False),
    ("feature_identification", "str",   "Sub-tipe fitur",      False),
    ("dim_classification",     "str",   "Klasifikasi dimensi", False),
    ("clock_position",         "str",   "Posisi jam (HH:MM)",  False),
    ("wall_thickness_mm",      "float", "Wall thickness (mm)", False),
    ("length_mm",              "float", "Panjang anomali (mm)",False),
    ("width_mm",               "float", "Lebar anomali (mm)",  False),
    ("depth_pct",              "float", "Kedalaman (% WT)",    False),
    ("surface_location",       "str",   "Internal/External",   False),
    ("erf_asme",               "float", "ERF ASME B31G",       False),
    ("psafe_bar",              "float", "P-safe ASME (bar)",   False),
    ("comment",                "str",   "Komentar vendor",     False),
    ("_source_file",           "str",   "[AUTO]",              False),
    ("_survey_year",           "int",   "[AUTO]",              False),
    ("_vendor",                "str",   "[AUTO]",              False),
    ("_import_timestamp",      "str",   "[AUTO]",              False),
]
UNIVERSAL_COLS = [f[0] for f in UNIVERSAL_SCHEMA]

FEATURE_TYPE_NORMALIZATION = {
    "valve": "Valve", "valv": "Valve", "ball valve": "Valve",
    "gate valve": "Valve", "block valve": "Valve",
    "weld": "Weld", "girth weld": "Weld", "seam weld": "Weld",
    "girth weld anomaly": "Weld", "longitudinal weld": "Weld",
    "longitudinal weld pipe": "Weld",
    "launcher end": "Weld", "chwt": "Weld",
    "valve pipe": "Weld",
    "bend": "Bend", "bend begin": "Bend", "bend end": "Bend",
    "elbow": "Bend", "bend start": "Bend",
    "anomaly": "Metal Loss", "metal loss": "Metal Loss",
    "corrosion": "Metal Loss", "external metal": "Metal Loss",
    "external metal loss": "Metal Loss",
    "internal metal loss": "Metal Loss", "anom": "Metal Loss",
    "cocl": "Metal Loss", "corr": "Metal Loss",
    "metal loss cluster": "Metal Loss", "cluster corrosion": "Metal Loss",
    "pipe mill anomaly": "Metal Loss",
    "dent": "Dent", "dent part": "Dent", "denp": "Dent",
    "mechanical damage": "Dent",
    "anode": "Anode", "anod": "Anode", "sacrificial anode": "Anode",
    "offtake": "Offtake", "offtake tap": "Offtake", "offtake tee": "Offtake",
    "offt": "Offtake", "off take": "Offtake",
    "comp": "Fitting", "full circle fitting": "Fitting",
    "repair shell": "Fitting", "fitting": "Fitting", "repa": "Fitting",
    "clamp": "Fitting", "sleeve begin": "Fitting", "sleeve end": "Fitting",
    "tee": "Tee", "field tee": "Tee",
    "flange": "Flange", "joint flange": "Flange",
    "support": "Support", "external support": "Support",
    "attachment": "Support",
    "marker": "Marker", "above ground marker": "Marker", "agm": "Marker",
    "data recording begin": "Recording", "data recording end": "Recording",
    "othe": "Other", "other": "Other", "pipeline fixture": "Other",
    "casing": "Other", "casing begin": "Other", "casing end": "Other",
    "joint insulation": "Other", "extraneous body": "Other",
}

SURFACE_NORMALIZATION = {
    "internal": "Internal", "int": "Internal", "midwall": "Internal",
    "mid-wall": "Internal", "mid wall": "Internal", "inside": "Internal",
    "external": "External", "ext": "External", "non-internal": "External",
    "non internal": "External", "outer": "External", "outside": "External",
}

ANOMALY_TYPES  = {"Metal Loss", "Dent"}
LANDMARK_TYPES = {"Valve", "Weld", "Tee"}


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 2 — DATACLASS
# ══════════════════════════════════════════════════════════════════════════════════

@dataclass
class Spool:
    """Representasi satu spool (segmen antara dua girth weld)."""
    spool_id: str
    us_weld_id: str
    ds_weld_id: str
    odo_us: float
    odo_ds: float

    @property
    def length(self) -> float:
        return self.odo_ds - self.odo_us


@dataclass
class SpoolPair:
    """Pasangan spool dari dua run yang sudah di-match pada Layer 2."""
    spool_r1: Spool
    spool_r2: Spool
    joint_len_r1: float
    joint_len_r2: float
    len_diff_pct: float
    consistent: bool


@dataclass
class MatchResult:
    """Hasil matching satu pasang anomali."""
    spool_id: str
    anom_id_r1: Optional[str]
    anom_id_r2: Optional[str]
    delta_odo: Optional[float]
    delta_clock: Optional[float]
    depth_r1: Optional[float]
    depth_r2: Optional[float]
    length_r1: Optional[float]
    length_r2: Optional[float]
    side: Optional[str]
    cost: float
    status: str = "UNVALIDATED"
    flag: str = ""
    depth_growth_pct: Optional[float] = None
    growth_rate_per_yr: Optional[float] = None
    erf: Optional[float] = None
    remaining_life_yr: Optional[float] = None


@dataclass
class AlignmentReport:
    """Laporan lengkap alignment pipeline."""
    n_valves_r1: int
    n_valves_r2: int
    n_welds_matched: int
    n_welds_unmatched_r1: int
    n_welds_unmatched_r2: int
    n_spool_pairs: int
    match_results: list[MatchResult] = field(default_factory=list)
    correction_factors: list[dict] = field(default_factory=list)
    weld_matching_summary: list[dict] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    _corrected_run1: Optional[dict] = None
    _corrected_run2: Optional[dict] = None
    _spool_pairs: list = field(default_factory=list)


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 3 — FUZZY COLUMN MATCHING
# ══════════════════════════════════════════════════════════════════════════════════

# Alias kolom: kunci = nama standar, nilai = daftar nama alternatif vendor
_COLUMN_ALIASES: dict[str, list[str]] = {
    "feature_id":             ["feature id", "anomaly id", "anom id", "id",
                               "feat id", "anomaly_id", "identification id",
                               "feature_no", "feature no", "ref", "reference",
                               "feature #", "feature"],
    "log_distance_m":         ["log dist", "log distance", "odometer", "odo",
                               "absolute distance", "abs dist", "distance",
                               "log_dist", "chainage", "kp", "log distance (m)",
                               "log_distance", "abs_dist", "distance_m",
                               "log dist (m)", "cumulative distance",
                               "abs distance (m)", "abs distance"],
    "joint_number":           ["joint no", "joint", "weld no", "pipe no",
                               "joint_no", "pipe joint", "jt no", "jt",
                               "joint number", "pipe number", "pipe_no",
                               "girth weld nr"],
    "dist_to_us_gw_m":        ["dist to us gw", "us weld dist", "distance to upstream",
                               "dist upstream", "dist us gw", "dist_us_gw",
                               "dist to upstream gw", "us gw dist",
                               "distance from upstream girth weld",
                               "dist to us girth weld", "dist us",
                               "dist u/s gw (m)", "dist u s gw (m)",
                               "dist u/s gw"],
    "joint_length_m":         ["joint length", "pipe length", "joint len",
                               "jt length", "pipe len", "joint_len",
                               "pipe_length", "spool length",
                               "joint length (m)"],
    "feature_type":           ["feature type", "type", "feat type",
                               "feature_type", "anomaly type", "anom type",
                               "component", "component type",
                               "feature type *"],
    "feature_identification": ["feature identification", "identification",
                               "feature ident", "feat ident", "sub type",
                               "sub-type", "subtype", "feature_ident",
                               "anomaly sub-type", "anomaly classification"],
    "dim_classification":     ["dim classification", "dimension", "classification",
                               "dim class", "dim_class", "size classification",
                               "dimensional classification",
                               "dimension classification"],
    "clock_position":         ["clock", "clock pos", "orientation", "o'clock",
                               "clock position", "oclock", "clock_pos",
                               "circumferential position", "circ position",
                               "circ pos", "clock (o'clock)",
                               "orientation o clock", "orientation o'clock"],
    "wall_thickness_mm":      ["wall thickness", "wt", "thickness", "wt mm",
                               "wall_thickness", "wt_mm", "nom wt",
                               "nominal wall thickness", "pipe wt",
                               "wall thickness (mm)"],
    "length_mm":              ["length", "axial length", "defect length",
                               "anomaly length", "len", "length mm",
                               "length_mm", "axial len", "anom length",
                               "axial extent", "defect len",
                               "axial length (mm)"],
    "width_mm":               ["width", "circ width", "defect width",
                               "anomaly width", "wid", "width mm",
                               "width_mm", "circumferential width",
                               "circ extent", "defect wid",
                               "width (mm)"],
    "depth_pct":              ["depth", "depth %", "peak depth", "max depth",
                               "depth pct", "depth_pct", "depth (% wt)",
                               "depth percent", "d/t", "depth (%wt)",
                               "depth % wt", "peak depth (%wt)",
                               "anomaly depth"],
    "surface_location":       ["surface", "surface location", "int/ext",
                               "location", "int ext", "surface_location",
                               "internal external", "surface loc"],
    "erf_asme":               ["erf", "erf asme", "erf b31g", "erf_asme",
                               "estimated repair factor", "repair factor",
                               "erf (asme b31g)", "erf asme b31g"],
    "psafe_bar":              ["psafe", "p safe", "safe pressure", "psafe bar",
                               "psafe_bar", "safe operating pressure",
                               "p-safe", "calculated safe pressure",
                               "psafe asme b31g (bar)", "psafe asme b31g"],
    "comment":                ["comment", "comments", "remark", "remarks",
                               "note", "notes", "description", "vendor comment"],
}


def _normalize_col_name(name: str) -> str:
    """Normalisasi nama kolom: lowercase, strip, hapus karakter spesial."""
    s = str(name).strip().lower()
    s = re.sub(r"[_\-./()\*#]+", " ", s)  # Hapus *, # juga
    s = re.sub(r"\s+", " ", s).strip()
    return s


def fuzzy_match_columns(raw_columns: list[str]) -> dict[str, str]:
    """
    Cocokkan kolom raw Excel ke skema universal menggunakan alias table.
    Return: dict mapping {raw_column_name -> universal_column_name}
    """
    mapping: dict[str, str] = {}
    used_universal: set[str] = set()

    for raw_col in raw_columns:
        norm = _normalize_col_name(raw_col)
        best_match: Optional[str] = None

        # Pass 1: exact match terhadap nama universal
        for uni_col in UNIVERSAL_COLS:
            if norm == uni_col.replace("_", " "):
                best_match = uni_col
                break

        # Pass 2: match terhadap alias
        if best_match is None:
            for uni_col, aliases in _COLUMN_ALIASES.items():
                if uni_col in used_universal:
                    continue
                for alias in aliases:
                    if norm == alias or norm.startswith(alias) or alias.startswith(norm):
                        best_match = uni_col
                        break
                if best_match:
                    break

        # Pass 3: substring match (partial)
        if best_match is None:
            for uni_col, aliases in _COLUMN_ALIASES.items():
                if uni_col in used_universal:
                    continue
                for alias in aliases:
                    if alias in norm or norm in alias:
                        best_match = uni_col
                        break
                if best_match:
                    break

        if best_match and best_match not in used_universal:
            mapping[raw_col] = best_match
            used_universal.add(best_match)

    return mapping


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 4 — LOADER: BACA EXCEL → UNIVERSAL DATAFRAME
# ══════════════════════════════════════════════════════════════════════════════════

def _clock_to_degrees(clock_str: str) -> Optional[float]:
    """Konversi posisi jam (HH:MM, HH:MM:SS, atau desimal) → derajat (0-360)."""
    if not clock_str or pd.isna(clock_str):
        return None
    s = str(clock_str).strip()
    # Format HH:MM:SS (dari Excel time)
    m3 = re.match(r"(\d{1,2}):(\d{2}):(\d{2})", s)
    if m3:
        h, mi = int(m3.group(1)), int(m3.group(2))
        total_min = h * 60 + mi
        return (total_min / 720.0) * 360.0
    # Format HH:MM
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        total_min = h * 60 + mi
        return (total_min / 720.0) * 360.0  # 12 jam = 720 menit = 360°
    # Format desimal (mis. 3.5)
    m2 = re.match(r"(\d{1,2}(?:\.\d+)?)", s)
    if m2:
        h = float(m2.group(1))
        return (h / 12.0) * 360.0
    return None


def _degrees_to_clock(deg: float) -> str:
    """Konversi derajat (0-360) → posisi jam HH:MM."""
    deg = deg % 360
    total_min = (deg / 360.0) * 720.0
    h = int(total_min // 60)
    mi = int(total_min % 60)
    if h == 0:
        h = 12
    return f"{h:02d}:{mi:02d}"


def _clock_delta_degrees(c1: str, c2: str) -> Optional[float]:
    """Hitung delta posisi jam dalam derajat (0-180), mempertimbangkan circular."""
    d1 = _clock_to_degrees(c1)
    d2 = _clock_to_degrees(c2)
    if d1 is None or d2 is None:
        return None
    diff = abs(d1 - d2) % 360
    return min(diff, 360 - diff)


def _clock_delta_hours(c1: str, c2: str) -> Optional[float]:
    """Hitung delta posisi jam dalam jam (0-6), mempertimbangkan circular."""
    deg = _clock_delta_degrees(c1, c2)
    if deg is None:
        return None
    return deg / 30.0  # 30° per jam


def _try_read_field_map(file_path: str) -> dict[str, str]:
    """
    Coba baca sheet FIELD_MAP untuk mendapatkan mapping kolom vendor → universal.
    Return: {vendor_column_name -> universal_field_name}
    """
    mapping = {}
    try:
        df_fm = pd.read_excel(file_path, sheet_name='FIELD_MAP', header=1)
        cols = list(df_fm.columns)
        # Cari kolom 'Universal Field' dan 'Nama Kolom di File Anda'
        uni_col_idx = None
        vendor_col_idx = None
        for i, c in enumerate(cols):
            cl = str(c).lower()
            if 'universal' in cl and 'field' in cl:
                uni_col_idx = i
            elif 'nama kolom' in cl or 'vendor column' in cl or 'file anda' in cl:
                vendor_col_idx = i
        if uni_col_idx is not None and vendor_col_idx is not None:
            for _, row in df_fm.iterrows():
                uni = row.iloc[uni_col_idx]
                vendor = row.iloc[vendor_col_idx]
                if pd.notna(uni) and pd.notna(vendor):
                    mapping[str(vendor).strip()] = str(uni).strip()
    except Exception:
        pass
    return mapping


def load_ili_universal(
    file_path: str,
    survey_year: Optional[int] = None,
    vendor: Optional[str] = None,
    sheet_name: int | str = 0,
) -> pd.DataFrame:
    """
    Baca file Excel ILI vendor ke format universal DataFrame.
    
    Langkah:
    1. Baca Excel, deteksi header row (coba sheet ILI_DATA dulu)
    2. Baca FIELD_MAP jika ada, untuk mapping kolom vendor
    3. Fuzzy-match kolom ke skema universal
    4. Normalisasi tipe fitur dan surface location
    5. Tambahkan metadata kolom (_source_file, _survey_year, dll)
    """
    print(f"\n{'='*70}")
    print(f"  LOADING: {os.path.basename(file_path)}")
    print(f"{'='*70}")

    # ── Deteksi sheet terbaik ──────────────────────────────────────────────
    try:
        xl = pd.ExcelFile(file_path)
        available_sheets = xl.sheet_names
        # Preferensi: ILI_DATA > sheet pertama
        if 'ILI_DATA' in available_sheets:
            sheet_name = 'ILI_DATA'
        elif sheet_name == 0 and len(available_sheets) > 0:
            sheet_name = available_sheets[0]
        print(f"  Sheet: {sheet_name} (available: {available_sheets})")
    except Exception:
        pass

    # ── Baca FIELD_MAP jika ada ────────────────────────────────────────────
    field_map = _try_read_field_map(file_path)
    if field_map:
        print(f"  FIELD_MAP ditemukan: {len(field_map)} mappings")

    # ── Baca Excel dengan deteksi header row terbaik ───────────────────────
    best_df = None
    best_header = 0
    best_map = {}
    best_score = -1

    for try_header in [0, 1, 2, 3, 4, 5]:
        try:
            df_temp = pd.read_excel(file_path, sheet_name=sheet_name, header=try_header)
            df_temp = df_temp.dropna(how='all').reset_index(drop=True)
            
            # Coba mapping kolom untuk header ini
            temp_map = {}
            remaining = []
            if field_map:
                for raw_c in df_temp.columns:
                    str_c = str(raw_c).strip()
                    if str_c in field_map:
                        temp_map[raw_c] = field_map[str_c]
                    else:
                        remaining.append(raw_c)
            else:
                remaining = list(df_temp.columns)
                
            if remaining:
                f_map = fuzzy_match_columns(remaining)
                for rc, uc in f_map.items():
                    temp_map[rc] = uc
            
            score = len(temp_map)
            # Berikan bobot ekstra jika log_distance_m dan feature_type ketemu
            if "log_distance_m" in temp_map.values():
                score += 5
            if "feature_type" in temp_map.values():
                score += 5
                
            if score > best_score:
                best_score = score
                best_header = try_header
                best_df = df_temp
                best_map = temp_map
                
        except Exception:
            continue

    if best_df is None or best_score == -1:
        raise ValueError(f"Gagal membaca file: {file_path}")

    df_raw = best_df
    header_row = best_header
    col_map = best_map

    print(f"  Header row: {header_row}, Shape: {df_raw.shape}")
    print(f"  Kolom raw: {list(df_raw.columns[:10])}...")
    print(f"  Kolom ter-mapping: {len(col_map)}/{len(UNIVERSAL_COLS)}")
    for raw_c, uni_c in sorted(col_map.items(), key=lambda x: str(x[1])):
        print(f"    {raw_c:40s} → {uni_c}")

    # ── Bangun DataFrame universal ─────────────────────────────────────────
    df = pd.DataFrame(index=df_raw.index)
    for uni_col in UNIVERSAL_COLS:
        # Cari kolom raw yang di-mapping ke uni_col
        raw_col = None
        for rc, uc in col_map.items():
            if uc == uni_col:
                raw_col = rc
                break
        if raw_col and raw_col in df_raw.columns:
            df[uni_col] = df_raw[raw_col]
        else:
            df[uni_col] = np.nan

    # ── Normalisasi feature_type ───────────────────────────────────────────
    def _norm_feature_type(val):
        if pd.isna(val):
            return "Other"
        s = str(val).strip().lower()
        return FEATURE_TYPE_NORMALIZATION.get(s, str(val).strip())

    df["feature_type"] = df["feature_type"].apply(_norm_feature_type)

    # ── Normalisasi surface_location ───────────────────────────────────────
    def _norm_surface(val):
        if pd.isna(val):
            return None
        s = str(val).strip().lower()
        return SURFACE_NORMALIZATION.get(s, str(val).strip())

    df["surface_location"] = df["surface_location"].apply(_norm_surface)

    # ── Konversi numerik ───────────────────────────────────────────────────
    numeric_cols = ["log_distance_m", "joint_number", "dist_to_us_gw_m",
                    "joint_length_m", "wall_thickness_mm", "length_mm",
                    "width_mm", "depth_pct", "erf_asme", "psafe_bar"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # ── Metadata ───────────────────────────────────────────────────────────
    df["_source_file"] = os.path.basename(file_path)
    df["_survey_year"] = survey_year if survey_year else _guess_year(file_path)
    df["_vendor"] = vendor if vendor else _guess_vendor(file_path)
    df["_import_timestamp"] = datetime.now().isoformat()

    # ── Hapus baris tanpa odometer ─────────────────────────────────────────
    n_before = len(df)
    df = df.dropna(subset=["log_distance_m"]).reset_index(drop=True)
    n_dropped = n_before - len(df)
    if n_dropped > 0:
        print(f"  ⚠ Hapus {n_dropped} baris tanpa odometer")

    # ── Sort by odometer ───────────────────────────────────────────────────
    df = df.sort_values("log_distance_m").reset_index(drop=True)

    # ── Generate feature_id jika kosong ────────────────────────────────────
    mask_no_id = df["feature_id"].isna() | (df["feature_id"].astype(str).str.strip() == "")
    if mask_no_id.any():
        for idx in df[mask_no_id].index:
            ft = df.loc[idx, "feature_type"]
            odo = df.loc[idx, "log_distance_m"]
            df.loc[idx, "feature_id"] = f"{ft}_{odo:.2f}_{idx}"

    print(f"  ✓ Loaded {len(df)} fitur")
    ft_counts = df["feature_type"].value_counts()
    for ft, cnt in ft_counts.items():
        print(f"    {ft:20s}: {cnt}")

    return df


def _guess_year(file_path: str) -> Optional[int]:
    """Coba tebak tahun survey dari nama file."""
    basename = os.path.basename(file_path)
    m = re.search(r"(20\d{2})", basename)
    return int(m.group(1)) if m else None


def _guess_vendor(file_path: str) -> Optional[str]:
    """Coba tebak vendor dari nama file."""
    basename = os.path.basename(file_path).lower()
    vendors = ["rosen", "tdw", "pigging", "baker", "ge", "lin",
               "eddyfi", "quest", "enduro", "pii", "bj"]
    for v in vendors:
        if v in basename:
            return v.upper()
    return None


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 5 — CONVERTER: DATAFRAME → RUN DICT
# ══════════════════════════════════════════════════════════════════════════════════

def df_to_run_dict(df: pd.DataFrame) -> dict:
    """
    Konversi universal DataFrame ke run dictionary:
    {
        "valves":    [{"id": ..., "odo": ...}, ...],
        "welds":     [{"id": ..., "odo": ..., "joint_num": ...}, ...],
        "anomalies": [{"id": ..., "odo": ..., "clock": ..., "depth": ..., ...}, ...],
        "all_features": [...]
    }
    Semua list sudah sorted by odometer.
    """
    run = {"valves": [], "welds": [], "anomalies": [], "all_features": []}

    for idx, row in df.iterrows():
        feat = {
            "id":           str(row.get("feature_id", f"F_{idx}")),
            "odo":          float(row.get("log_distance_m", 0)),
            "type":         str(row.get("feature_type", "Other")),
            "clock":        str(row.get("clock_position", "")) if pd.notna(row.get("clock_position")) else "",
            "depth":        float(row["depth_pct"]) if pd.notna(row.get("depth_pct")) else None,
            "length":       float(row["length_mm"]) if pd.notna(row.get("length_mm")) else None,
            "width":        float(row["width_mm"]) if pd.notna(row.get("width_mm")) else None,
            "wt":           float(row["wall_thickness_mm"]) if pd.notna(row.get("wall_thickness_mm")) else None,
            "surface":      str(row.get("surface_location", "")) if pd.notna(row.get("surface_location")) else "",
            "joint_num":    int(row["joint_number"]) if pd.notna(row.get("joint_number")) else None,
            "joint_len":    float(row["joint_length_m"]) if pd.notna(row.get("joint_length_m")) else None,
            "dist_us_gw":   float(row["dist_to_us_gw_m"]) if pd.notna(row.get("dist_to_us_gw_m")) else None,
            "erf":          float(row["erf_asme"]) if pd.notna(row.get("erf_asme")) else None,
            "psafe":        float(row["psafe_bar"]) if pd.notna(row.get("psafe_bar")) else None,
            "comment":      str(row.get("comment", "")) if pd.notna(row.get("comment")) else "",
            "ident":        str(row.get("feature_identification", "")) if pd.notna(row.get("feature_identification")) else "",
            "dim_class":    str(row.get("dim_classification", "")) if pd.notna(row.get("dim_classification")) else "",
        }
        run["all_features"].append(feat)

        ftype = feat["type"]
        if ftype == "Valve":
            run["valves"].append(feat)
        elif ftype == "Weld":
            run["welds"].append(feat)
        elif ftype in ANOMALY_TYPES:
            run["anomalies"].append(feat)

    # Pastikan sorted by odometer
    for key in ["valves", "welds", "anomalies", "all_features"]:
        run[key].sort(key=lambda x: x["odo"])

    print(f"\n  Run dict: {len(run['valves'])} valves, "
          f"{len(run['welds'])} welds, "
          f"{len(run['anomalies'])} anomalies, "
          f"{len(run['all_features'])} total features")

    return run


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 6 — LAYER 0: VALIDASI DATA
# ══════════════════════════════════════════════════════════════════════════════════

def layer0_validate(run: dict, label: str = "Run") -> list[str]:
    """
    Validasi integritas data sebelum alignment.

    Cek:
      1. Odometer monoton naik (sorted)
      2. Minimal ada 2 valve untuk landmark
      3. Minimal ada weld data
      4. Anomali punya depth & clock
      5. Duplikat feature_id
      6. Odometer gap besar (> 500m antara fitur berurutan)
    """
    print(f"\n{'─'*70}")
    print(f"  LAYER 0 — VALIDASI DATA [{label}]")
    print(f"{'─'*70}")

    errors: list[str] = []
    warnings_list: list[str] = []

    # ── Cek 1: Odometer monoton ────────────────────────────────────────────
    all_odos = [f["odo"] for f in run["all_features"]]
    for i in range(1, len(all_odos)):
        if all_odos[i] < all_odos[i-1]:
            errors.append(f"[L0] Odometer tidak monoton di index {i}: "
                         f"{all_odos[i-1]:.2f} → {all_odos[i]:.2f}")
            break

    # ── Cek 2: Minimal valve ───────────────────────────────────────────────
    n_valves = len(run["valves"])
    if n_valves < 2:
        warnings_list.append(f"[L0] Hanya {n_valves} valve ditemukan. "
                            f"Layer 1 correction mungkin terbatas.")
    print(f"  Valves: {n_valves}")

    # ── Cek 3: Minimal weld ───────────────────────────────────────────────
    n_welds = len(run["welds"])
    if n_welds < 2:
        errors.append(f"[L0] Hanya {n_welds} weld ditemukan. "
                     f"Tidak cukup untuk alignment.")
    print(f"  Welds: {n_welds}")

    # ── Cek 4: Anomali validasi ────────────────────────────────────────────
    n_anom = len(run["anomalies"])
    n_no_depth = sum(1 for a in run["anomalies"] if a["depth"] is None)
    n_no_clock = sum(1 for a in run["anomalies"] if not a["clock"])
    if n_no_depth > 0:
        warnings_list.append(f"[L0] {n_no_depth}/{n_anom} anomali tanpa depth")
    if n_no_clock > 0:
        warnings_list.append(f"[L0] {n_no_clock}/{n_anom} anomali tanpa clock position")
    print(f"  Anomalies: {n_anom} (no depth: {n_no_depth}, no clock: {n_no_clock})")

    # ── Cek 5: Duplikat ID ────────────────────────────────────────────────
    all_ids = [f["id"] for f in run["all_features"]]
    seen = set()
    dupes = set()
    for fid in all_ids:
        if fid in seen:
            dupes.add(fid)
        seen.add(fid)
    if dupes:
        warnings_list.append(f"[L0] {len(dupes)} duplikat feature_id: "
                            f"{list(dupes)[:5]}...")

    # ── Cek 6: Gap besar ──────────────────────────────────────────────────
    max_gap = 0
    for i in range(1, len(all_odos)):
        gap = all_odos[i] - all_odos[i-1]
        max_gap = max(max_gap, gap)
        if gap > 500:
            warnings_list.append(f"[L0] Gap besar {gap:.1f}m antara "
                                f"index {i-1} dan {i}")
    print(f"  Max gap antar fitur: {max_gap:.1f}m")

    # ── Report ─────────────────────────────────────────────────────────────
    for w in warnings_list:
        print(f"  ⚠ {w}")
    for e in errors:
        print(f"  ✗ {e}")
    if not errors:
        print(f"  ✓ Validasi {label} OK")

    return errors


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 7 — LAYER 1: VALVE-TO-VALVE ODOMETER CORRECTION
# ══════════════════════════════════════════════════════════════════════════════════

def _match_valves(valves_r1: list[dict], valves_r2: list[dict],
                  max_dist_m: float = 50.0) -> list[tuple[dict, dict]]:
    """
    Cocokkan valve berdasarkan urutan dan jarak odometer.
    Strategi: sequential matching — valve#1 R1 ↔ valve#1 R2, dst.
    Validasi: jarak antar valve pair harus < max_dist_m.
    """
    pairs = []
    n = min(len(valves_r1), len(valves_r2))
    for i in range(n):
        v1, v2 = valves_r1[i], valves_r2[i]
        dist = abs(v1["odo"] - v2["odo"])
        if dist < max_dist_m:
            pairs.append((v1, v2))
        else:
            print(f"    ⚠ Valve pair {i} terlalu jauh: {dist:.2f}m, skip")
    return pairs


def layer1_valve_correction(
    run1: dict, run2: dict
) -> tuple[dict, dict, list[dict]]:
    """
    Koreksi odometer menggunakan valve-to-valve landmark matching.

    Metode: Piecewise-linear interpolation antar valve pairs.
    Untuk fitur di antara dua valve pair, kita interpolasi linear
    faktor koreksi berdasarkan posisi relatif.

    Return:
        run1_corrected, run2_corrected, correction_factors
    """
    print(f"\n{'─'*70}")
    print(f"  LAYER 1 — VALVE-TO-VALVE ODOMETER CORRECTION")
    print(f"{'─'*70}")

    valve_pairs = _match_valves(run1["valves"], run2["valves"])
    n_pairs = len(valve_pairs)
    print(f"  Valve pairs matched: {n_pairs}")

    if n_pairs < 1:
        print("  ⚠ Tidak ada valve pair — skip koreksi")
        return copy.deepcopy(run1), copy.deepcopy(run2), []

    # ── Hitung correction factor di setiap valve pair ──────────────────────
    correction_factors = []
    for i, (v1, v2) in enumerate(valve_pairs):
        offset = v2["odo"] - v1["odo"]
        correction_factors.append({
            "pair_idx": i,
            "valve_r1_id": v1["id"],
            "valve_r2_id": v2["id"],
            "odo_r1": v1["odo"],
            "odo_r2": v2["odo"],
            "offset": offset,
        })
        print(f"    Pair {i}: R1={v1['odo']:.2f}m, R2={v2['odo']:.2f}m, "
              f"offset={offset:+.3f}m")

    # ── Bangun fungsi interpolasi ──────────────────────────────────────────
    cf_odos_r1 = [cf["odo_r1"] for cf in correction_factors]
    cf_offsets = [cf["offset"] for cf in correction_factors]

    def _interpolate_offset(odo_r1: float) -> float:
        """Interpolasi offset untuk odometer R1 tertentu."""
        if len(cf_odos_r1) == 1:
            return cf_offsets[0]
        # Clamp ke boundary
        if odo_r1 <= cf_odos_r1[0]:
            return cf_offsets[0]
        if odo_r1 >= cf_odos_r1[-1]:
            return cf_offsets[-1]
        # Cari segmen
        idx = bisect.bisect_right(cf_odos_r1, odo_r1) - 1
        idx = min(idx, len(cf_odos_r1) - 2)
        # Linear interpolation
        x0, x1 = cf_odos_r1[idx], cf_odos_r1[idx + 1]
        y0, y1 = cf_offsets[idx], cf_offsets[idx + 1]
        if x1 == x0:
            return y0
        t = (odo_r1 - x0) / (x1 - x0)
        return y0 + t * (y1 - y0)

    # ── Terapkan koreksi ke Run 1 (adjust R1 mendekati R2) ─────────────────
    run1_c = copy.deepcopy(run1)
    for key in ["all_features", "valves", "welds", "anomalies"]:
        for feat in run1_c[key]:
            offset = _interpolate_offset(feat["odo"])
            feat["odo_original"] = feat["odo"]
            feat["odo"] = feat["odo"] + offset

    # Run 2 tidak dikoreksi
    run2_c = copy.deepcopy(run2)
    for key in ["all_features", "valves", "welds", "anomalies"]:
        for feat in run2_c[key]:
            feat["odo_original"] = feat["odo"]

    print(f"  ✓ Koreksi diterapkan ke {len(run1_c['all_features'])} fitur R1")
    return run1_c, run2_c, correction_factors


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 8 — LAYER 2: WELD-TO-WELD SEQUENTIAL MATCHING (TWO-POINTER)
# ══════════════════════════════════════════════════════════════════════════════════

def layer2_weld_matching(
    run1: dict, run2: dict,
    max_weld_dist_m: float = 3.0,
    joint_len_tol_pct: float = JOINT_LEN_CONSISTENCY_PCT,
) -> tuple[list[SpoolPair], list[dict], list[dict], list[dict]]:
    """
    Two-pointer weld matching dengan validasi joint length consistency.

    Algoritma:
      1. Two-pointer scan: untuk setiap weld R1, cari weld R2 terdekat
         dalam window ±max_weld_dist_m
      2. Validasi: joint length antar weld pair berurutan harus konsisten
         (selisih < joint_len_tol_pct)
      3. Bangun spool pairs dari weld pairs yang valid

    Return:
        spool_pairs, weld_match_summary, unmatched_r1, unmatched_r2
    """
    print(f"\n{'─'*70}")
    print(f"  LAYER 2 — WELD-TO-WELD SEQUENTIAL MATCHING")
    print(f"{'─'*70}")

    welds_r1 = run1["welds"]
    welds_r2 = run2["welds"]
    print(f"  Welds R1: {len(welds_r1)}, R2: {len(welds_r2)}")

    # ── Two-pointer matching ───────────────────────────────────────────────
    matched_pairs: list[tuple[dict, dict]] = []
    used_r2: set[int] = set()
    unmatched_r1_list: list[dict] = []

    j_start = 0  # pointer R2
    for i, w1 in enumerate(welds_r1):
        best_j = None
        best_dist = INF

        # Scan R2 dari j_start
        for j in range(j_start, len(welds_r2)):
            if j in used_r2:
                continue
            dist = abs(w1["odo"] - welds_r2[j]["odo"])
            if dist > max_weld_dist_m * 3:
                # Terlalu jauh ke kanan, stop scan
                if welds_r2[j]["odo"] > w1["odo"] + max_weld_dist_m * 3:
                    break
                continue
            if dist < best_dist:
                best_dist = dist
                best_j = j

        if best_j is not None and best_dist <= max_weld_dist_m:
            matched_pairs.append((w1, welds_r2[best_j]))
            used_r2.add(best_j)
            # Advance j_start
            j_start = max(j_start, best_j)
        else:
            unmatched_r1_list.append({
                "weld_id": w1["id"], "odo": w1["odo"], "run": "R1"
            })

    # Weld R2 yang tidak ter-match
    unmatched_r2_list = []
    for j, w2 in enumerate(welds_r2):
        if j not in used_r2:
            unmatched_r2_list.append({
                "weld_id": w2["id"], "odo": w2["odo"], "run": "R2"
            })

    print(f"  Matched weld pairs: {len(matched_pairs)}")
    print(f"  Unmatched R1: {len(unmatched_r1_list)}, R2: {len(unmatched_r2_list)}")

    # ── Bangun spool pairs ─────────────────────────────────────────────────
    spool_pairs: list[SpoolPair] = []
    weld_summary: list[dict] = []

    for k in range(len(matched_pairs)):
        w1, w2 = matched_pairs[k]
        weld_summary.append({
            "pair_idx":  k,
            "weld_r1":   w1["id"],
            "weld_r2":   w2["id"],
            "odo_r1":    w1["odo"],
            "odo_r2":    w2["odo"],
            "delta_odo": abs(w1["odo"] - w2["odo"]),
        })

    for k in range(len(matched_pairs) - 1):
        w1_us, w2_us = matched_pairs[k]
        w1_ds, w2_ds = matched_pairs[k + 1]

        spool_r1 = Spool(
            spool_id=f"SP_R1_{k}",
            us_weld_id=w1_us["id"],
            ds_weld_id=w1_ds["id"],
            odo_us=w1_us["odo"],
            odo_ds=w1_ds["odo"],
        )
        spool_r2 = Spool(
            spool_id=f"SP_R2_{k}",
            us_weld_id=w2_us["id"],
            ds_weld_id=w2_ds["id"],
            odo_us=w2_us["odo"],
            odo_ds=w2_ds["odo"],
        )

        len_r1 = spool_r1.length
        len_r2 = spool_r2.length
        avg_len = (len_r1 + len_r2) / 2.0 if (len_r1 + len_r2) > 0 else 1.0
        len_diff_pct = abs(len_r1 - len_r2) / avg_len

        consistent = len_diff_pct <= joint_len_tol_pct

        sp = SpoolPair(
            spool_r1=spool_r1,
            spool_r2=spool_r2,
            joint_len_r1=len_r1,
            joint_len_r2=len_r2,
            len_diff_pct=len_diff_pct,
            consistent=consistent,
        )
        spool_pairs.append(sp)

    n_consistent = sum(1 for sp in spool_pairs if sp.consistent)
    print(f"  Spool pairs: {len(spool_pairs)} "
          f"(consistent: {n_consistent}, inconsistent: {len(spool_pairs) - n_consistent})")

    return spool_pairs, weld_summary, unmatched_r1_list, unmatched_r2_list


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 9 — LAYER 3: HUNGARIAN PER SPOOL (ANOMALY MATCHING)
# ══════════════════════════════════════════════════════════════════════════════════

def _get_anomalies_in_spool(anomalies: list[dict], spool: Spool) -> list[dict]:
    """Ambil anomali yang berada di dalam spool (odo_us <= odo <= odo_ds)."""
    return [a for a in anomalies if spool.odo_us <= a["odo"] <= spool.odo_ds]


def _compute_matching_cost(a1: dict, a2: dict, spool_len: float) -> float:
    """
    Hitung cost matching antara dua anomali.

    Komponen cost:
      - Odometer distance (dinormalisasi terhadap panjang spool)
      - Clock position difference
      - Depth difference

    Hard gate:
      - Clock > CLOCK_HARD_GATE_HR (1.5 jam) → INF
      - Surface mismatch (INT vs EXT) → INF
    """
    # ── Hard gate: surface mismatch ────────────────────────────────────────
    s1 = a1.get("surface", "").strip()
    s2 = a2.get("surface", "").strip()
    if s1 and s2 and s1 != s2:
        return INF

    # ── Komponen 1: Odometer distance ──────────────────────────────────────
    odo_diff = abs(a1["odo"] - a2["odo"])
    norm_spool_len = max(spool_len, 1.0)
    odo_cost = odo_diff / norm_spool_len

    # ── Komponen 2: Clock difference ───────────────────────────────────────
    clock_diff_hr = _clock_delta_hours(a1.get("clock", ""), a2.get("clock", ""))
    if clock_diff_hr is not None:
        if clock_diff_hr > CLOCK_HARD_GATE_HR:
            return INF
        clock_cost = clock_diff_hr / 6.0  # Normalisasi ke [0,1]
    else:
        clock_cost = 0.5  # Penalty jika clock tidak tersedia

    # ── Komponen 3: Depth difference ───────────────────────────────────────
    d1 = a1.get("depth")
    d2 = a2.get("depth")
    if d1 is not None and d2 is not None:
        depth_cost = abs(d1 - d2) / 100.0  # Normalisasi ke [0,1]
    else:
        depth_cost = 0.3  # Penalty jika depth tidak tersedia

    # ── Weighted sum ───────────────────────────────────────────────────────
    total_cost = W_ODO * odo_cost + W_CLK * clock_cost + W_DEP * depth_cost
    return total_cost


def hungarian_per_spool(
    anom_r1: list[dict],
    anom_r2: list[dict],
    spool_pair: SpoolPair,
) -> list[MatchResult]:
    """
    Hungarian algorithm per spool untuk anomaly matching.

    1. Bangun cost matrix (n_r1 × n_r2)
    2. Tambah dummy rows/cols untuk handle unmatched
    3. Solve menggunakan scipy linear_sum_assignment
    4. Hasilkan MatchResult untuk setiap pasangan
    """
    spool_id = spool_pair.spool_r1.spool_id
    avg_spool_len = (spool_pair.spool_r1.length + spool_pair.spool_r2.length) / 2

    n1 = len(anom_r1)
    n2 = len(anom_r2)

    if n1 == 0 and n2 == 0:
        return []

    results: list[MatchResult] = []

    # ── Kasus: salah satu kosong ───────────────────────────────────────────
    if n1 == 0:
        for a2 in anom_r2:
            results.append(MatchResult(
                spool_id=spool_id, anom_id_r1=None, anom_id_r2=a2["id"],
                delta_odo=None, delta_clock=None,
                depth_r1=None, depth_r2=a2.get("depth"),
                length_r1=None, length_r2=a2.get("length"),
                side=a2.get("surface", ""), cost=INF,
                status="NEW_IN_R2",
            ))
        return results

    if n2 == 0:
        for a1 in anom_r1:
            results.append(MatchResult(
                spool_id=spool_id, anom_id_r1=a1["id"], anom_id_r2=None,
                delta_odo=None, delta_clock=None,
                depth_r1=a1.get("depth"), depth_r2=None,
                length_r1=a1.get("length"), length_r2=None,
                side=a1.get("surface", ""), cost=INF,
                status="NDF",
            ))
        return results

    # ── Bangun cost matrix ─────────────────────────────────────────────────
    # Ukuran: (n1 + n2) × (n1 + n2) — padded supaya square
    dim = n1 + n2
    cost_matrix = np.full((dim, dim), INF / 2)

    # Real costs (top-left n1×n2)
    for i in range(n1):
        for j in range(n2):
            cost_matrix[i, j] = _compute_matching_cost(
                anom_r1[i], anom_r2[j], avg_spool_len
            )

    # Dummy costs: izinkan unmatched (bottom-left dan top-right)
    dummy_cost = 0.8  # Threshold: lebih baik unmatched daripada bad match
    for i in range(n1):
        cost_matrix[i, n2 + i] = dummy_cost  # R1 unmatched (NDF)
    for j in range(n2):
        cost_matrix[n1 + j, j] = dummy_cost  # R2 unmatched (NEW)

    # Fill sisa dengan INF
    # (sudah di-fill di atas dengan INF/2, tapi corner cells harus 0)
    for k in range(min(n1, n2), dim):
        for l in range(min(n1, n2), dim):
            if k >= n1 and l >= n2:
                cost_matrix[k, l] = 0  # Dummy-to-dummy = free

    # ── Solve Hungarian ────────────────────────────────────────────────────
    row_ind, col_ind = linear_sum_assignment(cost_matrix)

    matched_r1: set[int] = set()
    matched_r2: set[int] = set()

    for r, c in zip(row_ind, col_ind):
        if r < n1 and c < n2:
            # Real match
            a1 = anom_r1[r]
            a2 = anom_r2[c]
            cost_val = cost_matrix[r, c]

            if cost_val >= INF / 2:
                # Cost terlalu tinggi, treat as unmatched
                continue

            delta_odo = abs(a1["odo"] - a2["odo"])
            delta_clock = _clock_delta_hours(a1.get("clock", ""), a2.get("clock", ""))

            results.append(MatchResult(
                spool_id=spool_id,
                anom_id_r1=a1["id"],
                anom_id_r2=a2["id"],
                delta_odo=delta_odo,
                delta_clock=delta_clock,
                depth_r1=a1.get("depth"),
                depth_r2=a2.get("depth"),
                length_r1=a1.get("length"),
                length_r2=a2.get("length"),
                side=a1.get("surface", "") or a2.get("surface", ""),
                cost=cost_val,
                status="MATCHED",
            ))
            matched_r1.add(r)
            matched_r2.add(c)

    # ── Unmatched R1 → NDF ─────────────────────────────────────────────────
    for i in range(n1):
        if i not in matched_r1:
            a1 = anom_r1[i]
            results.append(MatchResult(
                spool_id=spool_id, anom_id_r1=a1["id"], anom_id_r2=None,
                delta_odo=None, delta_clock=None,
                depth_r1=a1.get("depth"), depth_r2=None,
                length_r1=a1.get("length"), length_r2=None,
                side=a1.get("surface", ""), cost=INF,
                status="NDF",
            ))

    # ── Unmatched R2 → NEW_IN_R2 ──────────────────────────────────────────
    for j in range(n2):
        if j not in matched_r2:
            a2 = anom_r2[j]
            results.append(MatchResult(
                spool_id=spool_id, anom_id_r1=None, anom_id_r2=a2["id"],
                delta_odo=None, delta_clock=None,
                depth_r1=None, depth_r2=a2.get("depth"),
                length_r1=None, length_r2=a2.get("length"),
                side=a2.get("surface", ""), cost=INF,
                status="NEW_IN_R2",
            ))

    return results


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 10 — LAYER 3 ORCHESTRATION: ANOMALY MATCHING (per spool iteration)
# ══════════════════════════════════════════════════════════════════════════════════

def layer3_anomaly_matching(
    run1: dict, run2: dict,
    spool_pairs: list[SpoolPair],
) -> list[MatchResult]:
    """
    Orkestrasi Layer 3: iterasi setiap spool pair dan jalankan Hungarian matching.

    Anomali yang TIDAK berada di dalam spool pair manapun akan dilaporkan
    sebagai NDF (dari R1) atau NEW_IN_R2 (dari R2).

    Return: list[MatchResult] gabungan dari semua spool.
    """
    print(f"\n{'─'*70}")
    print(f"  LAYER 3 — ANOMALY MATCHING (HUNGARIAN PER SPOOL)")
    print(f"{'─'*70}")

    all_results: list[MatchResult] = []

    # Track anomali yang sudah masuk ke spool
    covered_r1_ids: set[str] = set()
    covered_r2_ids: set[str] = set()

    n_matched_total = 0
    n_ndf_total = 0
    n_new_total = 0

    for idx, sp in enumerate(spool_pairs):
        # ── Ambil anomali dalam spool ──────────────────────────────────────
        anom_r1 = _get_anomalies_in_spool(run1["anomalies"], sp.spool_r1)
        anom_r2 = _get_anomalies_in_spool(run2["anomalies"], sp.spool_r2)

        # ── Filter anomali yang belum di-cover ─────────────────────────────
        anom_r1 = [a for a in anom_r1 if a["id"] not in covered_r1_ids]
        anom_r2 = [a for a in anom_r2 if a["id"] not in covered_r2_ids]

        if not anom_r1 and not anom_r2:
            continue

        # ── Jalankan Hungarian per spool ───────────────────────────────────
        spool_results = hungarian_per_spool(anom_r1, anom_r2, sp)

        # ── Update tracking ────────────────────────────────────────────────
        for mr in spool_results:
            if mr.anom_id_r1:
                covered_r1_ids.add(mr.anom_id_r1)
            if mr.anom_id_r2:
                covered_r2_ids.add(mr.anom_id_r2)

            if mr.status == "MATCHED":
                n_matched_total += 1
            elif mr.status == "NDF":
                n_ndf_total += 1
            elif mr.status == "NEW_IN_R2":
                n_new_total += 1

        all_results.extend(spool_results)

        if len(anom_r1) > 0 or len(anom_r2) > 0:
            n_m = sum(1 for r in spool_results if r.status == "MATCHED")
            print(f"    Spool {idx:4d}: R1={len(anom_r1):3d}, R2={len(anom_r2):3d}, "
                  f"matched={n_m}")

    # ── Anomali TIDAK dalam spool pair manapun → NDF / NEW ─────────────────
    orphan_r1 = [a for a in run1["anomalies"] if a["id"] not in covered_r1_ids]
    orphan_r2 = [a for a in run2["anomalies"] if a["id"] not in covered_r2_ids]

    for a1 in orphan_r1:
        all_results.append(MatchResult(
            spool_id="NO_SPOOL",
            anom_id_r1=a1["id"], anom_id_r2=None,
            delta_odo=None, delta_clock=None,
            depth_r1=a1.get("depth"), depth_r2=None,
            length_r1=a1.get("length"), length_r2=None,
            side=a1.get("surface", ""), cost=INF,
            status="NDF",
            flag="outside_spool",
        ))
        n_ndf_total += 1

    for a2 in orphan_r2:
        all_results.append(MatchResult(
            spool_id="NO_SPOOL",
            anom_id_r1=None, anom_id_r2=a2["id"],
            delta_odo=None, delta_clock=None,
            depth_r1=None, depth_r2=a2.get("depth"),
            length_r1=None, length_r2=a2.get("length"),
            side=a2.get("surface", ""), cost=INF,
            status="NEW_IN_R2",
            flag="outside_spool",
        ))
        n_new_total += 1

    # ── Summary ────────────────────────────────────────────────────────────
    print(f"\n  ┌─────────────────────────────────────────────┐")
    print(f"  │  LAYER 3 SUMMARY                             │")
    print(f"  ├─────────────────────────────────────────────┤")
    print(f"  │  Total anomali R1:    {len(run1['anomalies']):6d}                 │")
    print(f"  │  Total anomali R2:    {len(run2['anomalies']):6d}                 │")
    print(f"  │  MATCHED:             {n_matched_total:6d}                 │")
    print(f"  │  NDF (hilang di R2):  {n_ndf_total:6d}                 │")
    print(f"  │  NEW (baru di R2):    {n_new_total:6d}                 │")
    print(f"  │  Orphan R1:           {len(orphan_r1):6d}                 │")
    print(f"  │  Orphan R2:           {len(orphan_r2):6d}                 │")
    print(f"  └─────────────────────────────────────────────┘")

    return all_results


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 11 — LAYER 4: GROWTH VALIDATION
# ══════════════════════════════════════════════════════════════════════════════════

def _calc_asme_b31g(
    depth_pct: float,
    length_mm: float,
    wt_mm: float,
    od_mm: float,
    smys_mpa: float,
    maop_bar: float,
) -> tuple[float, float, float]:
    """
    Hitung ERF, P-safe, dan remaining strength menggunakan Modified ASME B31G.

    Parameter:
        depth_pct:  kedalaman defek (% WT)
        length_mm:  panjang defek (mm)
        wt_mm:      wall thickness (mm)
        od_mm:      outside diameter (mm)
        smys_mpa:   Specified Minimum Yield Strength (MPa)
        maop_bar:   Maximum Allowable Operating Pressure (bar)

    Return:
        (erf, psafe_bar, remaining_strength_ratio)
    """
    # Konversi depth dari %WT ke mm
    d_mm = (depth_pct / 100.0) * wt_mm
    d_over_t = d_mm / wt_mm  # = depth_pct / 100

    # Panjang defek
    L = max(length_mm, 1.0)

    # Folias factor M (Modified B31G)
    L_sq = L * L
    od_wt = od_mm * wt_mm
    ratio = L_sq / od_wt

    if ratio <= 50.0:
        M = math.sqrt(1.0 + 0.6275 * ratio - 0.003375 * ratio * ratio)
    else:
        # Untuk defek sangat panjang, gunakan formulasi alternatif
        M = 0.032 * ratio + 3.3

    # Tekanan aman (P-safe) — Modified B31G
    # P_safe = (2 * SMYS * WT / OD) * (1 - (2/3)*(d/t)) / (1 - (2/3)*(d/t)/M)
    numerator = 1.0 - (2.0 / 3.0) * d_over_t
    denominator = 1.0 - (2.0 / 3.0) * d_over_t / M

    if denominator <= 0 or numerator <= 0:
        # Defek terlalu dalam, struktur gagal
        return INF, 0.0, 0.0

    # SMYS dalam MPa, konversi ke bar: 1 MPa = 10 bar
    p_flow_mpa = (2.0 * smys_mpa * wt_mm / od_mm) * (numerator / denominator)
    p_safe_bar = p_flow_mpa * 10.0  # MPa → bar

    # ERF = MAOP / P_safe
    if p_safe_bar > 0:
        erf = maop_bar / p_safe_bar
    else:
        erf = INF

    # Remaining strength ratio
    rs_ratio = numerator / denominator

    return erf, p_safe_bar, rs_ratio


def layer4_growth_validation(
    results: list[MatchResult],
    years_between: float,
    wt_mm: float = DEFAULT_WT_MM,
    od_mm: float = DEFAULT_OD_MM,
    smys_mpa: float = DEFAULT_SMYS_MPA,
    maop_bar: float = DEFAULT_MAOP_BAR,
) -> list[MatchResult]:
    """
    Layer 4: Validasi pertumbuhan anomali untuk setiap pasangan MATCHED.

    Constraint checks:
      C1: delta_odo within spool (dijamin oleh Layer 3)
      C2: Clock consistent (≤ 1.5 jam = 45°)
      C3: INT/EXT sama (dijamin oleh Layer 3 hard gate)
      C4: Depth tidak turun > C4_DEPTH_TOLERANCE_PCT_WT (2% WT)
      C5: Length tidak turun > 10% (C5_LENGTH_DECREASE_RATIO = 0.9)
      C6: Growth rate ≤ C6_MAX_GROWTH_RATE_PER_YR (1.5 %wt/yr)

    Klasifikasi:
      ACTIVE_CORROSION:  depth naik DAN growth_rate > 0
      STABLE:            perubahan depth dalam toleransi
      SUSPECT_MATCH:     constraint C4/C5 gagal signifikan

    Untuk ACTIVE_CORROSION, hitung:
      - ERF (Modified ASME B31G)
      - Remaining life = (allowable_depth - current_depth) / growth_rate
    """
    print(f"\n{'─'*70}")
    print(f"  LAYER 4 — GROWTH VALIDATION")
    print(f"{'─'*70}")
    print(f"  Years between surveys: {years_between:.1f}")
    print(f"  Pipeline: OD={od_mm}mm, WT={wt_mm}mm, SMYS={smys_mpa}MPa, "
          f"MAOP={maop_bar}bar")

    n_active = 0
    n_stable = 0
    n_suspect = 0
    n_skipped = 0

    # Kedalaman maksimum yang diizinkan (80% WT secara umum)
    allowable_depth_pct = 80.0

    for mr in results:
        # Hanya proses MATCHED pairs
        if mr.status != "MATCHED":
            continue

        depth_r1 = mr.depth_r1
        depth_r2 = mr.depth_r2
        length_r1 = mr.length_r1
        length_r2 = mr.length_r2

        # Skip jika data tidak lengkap
        if depth_r1 is None or depth_r2 is None:
            mr.status = "MATCHED"
            mr.flag = "incomplete_data"
            n_skipped += 1
            continue

        # ── Hitung pertumbuhan depth ───────────────────────────────────────
        depth_change = depth_r2 - depth_r1  # positif = tumbuh
        mr.depth_growth_pct = depth_change

        if years_between > 0:
            mr.growth_rate_per_yr = depth_change / years_between
        else:
            mr.growth_rate_per_yr = 0.0

        # ── C2: Clock consistency ──────────────────────────────────────────
        clock_ok = True
        if mr.delta_clock is not None:
            if mr.delta_clock > CLOCK_HARD_GATE_HR:
                clock_ok = False

        # ── C4: Depth decrease check ──────────────────────────────────────
        c4_ok = True
        if depth_change < -C4_DEPTH_TOLERANCE_PCT_WT:
            c4_ok = False  # Depth turun lebih dari toleransi

        # ── C5: Length decrease check ──────────────────────────────────────
        c5_ok = True
        if length_r1 is not None and length_r2 is not None:
            if length_r1 > 0:
                length_ratio = length_r2 / length_r1
                if length_ratio < C5_LENGTH_DECREASE_RATIO:
                    c5_ok = False  # Length turun > 10%

        # ── C6: Growth rate check ──────────────────────────────────────────
        c6_ok = True
        if mr.growth_rate_per_yr is not None:
            if mr.growth_rate_per_yr > C6_MAX_GROWTH_RATE_PER_YR:
                c6_ok = False  # Growth rate terlalu tinggi

        # ── Klasifikasi ───────────────────────────────────────────────────
        flags = []
        if not c4_ok:
            flags.append("C4_depth_decrease")
        if not c5_ok:
            flags.append("C5_length_decrease")
        if not c6_ok:
            flags.append("C6_high_growth")
        if not clock_ok:
            flags.append("C2_clock_shift")

        if not c4_ok or not c5_ok:
            # Hard constraint gagal → suspect match
            mr.status = "SUSPECT_MATCH"
            mr.flag = "; ".join(flags)
            n_suspect += 1
        elif depth_change > C4_DEPTH_TOLERANCE_PCT_WT and mr.growth_rate_per_yr > 0:
            # Depth naik signifikan → aktif korosi
            mr.status = "ACTIVE_CORROSION"
            if flags:
                mr.flag = "; ".join(flags)
            n_active += 1

            # ── Hitung ERF & remaining life ────────────────────────────────
            use_length = length_r2 if length_r2 is not None else 50.0
            erf, psafe, _ = _calc_asme_b31g(
                depth_pct=depth_r2,
                length_mm=use_length,
                wt_mm=wt_mm,
                od_mm=od_mm,
                smys_mpa=smys_mpa,
                maop_bar=maop_bar,
            )
            mr.erf = round(erf, 4) if erf < INF else None
            if mr.growth_rate_per_yr > 0:
                remaining_depth = allowable_depth_pct - depth_r2
                if remaining_depth > 0:
                    mr.remaining_life_yr = round(
                        remaining_depth / mr.growth_rate_per_yr, 1
                    )
                else:
                    mr.remaining_life_yr = 0.0
        else:
            # Perubahan dalam toleransi → stabil
            mr.status = "STABLE"
            if flags:
                mr.flag = "; ".join(flags)
            n_stable += 1

    # ── Summary ────────────────────────────────────────────────────────────
    print(f"\n  ┌─────────────────────────────────────────────┐")
    print(f"  │  LAYER 4 SUMMARY                             │")
    print(f"  ├─────────────────────────────────────────────┤")
    print(f"  │  ACTIVE_CORROSION:     {n_active:6d}                 │")
    print(f"  │  STABLE:               {n_stable:6d}                 │")
    print(f"  │  SUSPECT_MATCH:        {n_suspect:6d}                 │")
    print(f"  │  Skipped (no data):    {n_skipped:6d}                 │")
    print(f"  └─────────────────────────────────────────────┘")

    # ERF warnings
    critical_erf = [mr for mr in results
                    if mr.erf is not None and mr.erf >= 1.0]
    if critical_erf:
        print(f"\n  ⚠ {len(critical_erf)} anomali dengan ERF ≥ 1.0 (KRITIS):")
        for mr in critical_erf[:10]:
            print(f"    {mr.anom_id_r2}: ERF={mr.erf:.3f}, "
                  f"depth={mr.depth_r2:.1f}%, "
                  f"remaining_life={mr.remaining_life_yr}yr")

    return results


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 12 — EXCEL ENGINE (Production Multi-Sheet)
# ══════════════════════════════════════════════════════════════════════════════════

# ── Konstanta style ────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79",
                           fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center",
                          wrap_text=True)
_LEFT_ALIGN = Alignment(horizontal="left", vertical="center",
                        wrap_text=True)

# Row fills berdasarkan status
_FILL_ACTIVE   = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                             fill_type="solid")  # Light red
_FILL_STABLE   = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                             fill_type="solid")  # Light green
_FILL_SUSPECT  = PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                             fill_type="solid")  # Yellow
_FILL_NDF      = PatternFill(start_color="FCD5B4", end_color="FCD5B4",
                             fill_type="solid")  # Orange
_FILL_NEW      = PatternFill(start_color="BDD7EE", end_color="BDD7EE",
                             fill_type="solid")  # Light blue

_STATUS_FILLS = {
    "ACTIVE_CORROSION": _FILL_ACTIVE,
    "STABLE":           _FILL_STABLE,
    "SUSPECT_MATCH":    _FILL_SUSPECT,
    "NDF":              _FILL_NDF,
    "NEW_IN_R2":        _FILL_NEW,
}


def _style_header_row(ws, n_cols: int):
    """Terapkan style header ke row 1."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER_ALIGN


def _auto_column_width(ws, min_width: int = 10, max_width: int = 40):
    """Otomatis atur lebar kolom berdasarkan konten."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                max_len = max(max_len, cell_len)
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _apply_row_fill(ws, status_col_idx: int, start_row: int = 2):
    """Terapkan fill baris berdasarkan kolom status."""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        status_cell = row[status_col_idx - 1]  # 0-indexed dalam tuple
        status_val = str(status_cell.value) if status_cell.value else ""
        fill = _STATUS_FILLS.get(status_val)
        if fill:
            for cell in row:
                cell.fill = fill


def _write_df_to_sheet(ws, df: pd.DataFrame, apply_status_fill: bool = True):
    """Tulis DataFrame ke worksheet dengan formatting."""
    # Header
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    _style_header_row(ws, len(df.columns))

    # Data
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(df.columns, 1):
            val = row[col_name]
            # Handle NaN/None
            if pd.isna(val) if isinstance(val, (float, np.floating)) else val is None:
                cell_val = ""
            elif isinstance(val, float):
                cell_val = round(val, 4)
            else:
                cell_val = val
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
            cell.border = _THIN_BORDER
            cell.alignment = _LEFT_ALIGN

    # Status fill
    if apply_status_fill and "status" in df.columns:
        status_col_idx = list(df.columns).index("status") + 1
        _apply_row_fill(ws, status_col_idx)

    # Auto width & freeze
    _auto_column_width(ws)
    ws.freeze_panes = "A2"


def write_excel_report(
    report: AlignmentReport,
    output_path: str,
    df_r1: Optional[pd.DataFrame] = None,
    df_r2: Optional[pd.DataFrame] = None,
):
    """
    Tulis laporan alignment ke file Excel multi-sheet.

    Sheets:
      1. Summary        — statistik keseluruhan, ringkasan per layer
      2. Comparison      — semua pasangan matched, kolom R1/R2 side by side
      3. Only_NEW        — anomali hanya di R2 (NEW_IN_R2)
      4. Only_OLD        — anomali hanya di R1 (NDF)
      5. Physics_Valid   — hasil Layer 4 dengan klasifikasi growth
      6. Weld_Matching   — detail weld matching Layer 2
      7. Valve_Correction— faktor koreksi Layer 1
    """
    print(f"\n{'─'*70}")
    print(f"  EXCEL ENGINE — WRITING REPORT")
    print(f"{'─'*70}")

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════
    #  Sheet 1: Summary
    # ══════════════════════════════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "Summary"

    summary_data = [
        ("ILI Alignment Report", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("PIPELINE DATA", ""),
        ("Valves R1", report.n_valves_r1),
        ("Valves R2", report.n_valves_r2),
        ("", ""),
        ("LAYER 2 — WELD MATCHING", ""),
        ("Welds Matched", report.n_welds_matched),
        ("Welds Unmatched R1", report.n_welds_unmatched_r1),
        ("Welds Unmatched R2", report.n_welds_unmatched_r2),
        ("Spool Pairs", report.n_spool_pairs),
        ("", ""),
        ("LAYER 3 — ANOMALY MATCHING", ""),
    ]

    # Hitung statistik dari results
    n_matched = sum(1 for r in report.match_results if r.status == "MATCHED"
                    or r.status in ("ACTIVE_CORROSION", "STABLE", "SUSPECT_MATCH"))
    n_ndf = sum(1 for r in report.match_results if r.status == "NDF")
    n_new = sum(1 for r in report.match_results if r.status == "NEW_IN_R2")
    n_active = sum(1 for r in report.match_results
                   if r.status == "ACTIVE_CORROSION")
    n_stable = sum(1 for r in report.match_results if r.status == "STABLE")
    n_suspect = sum(1 for r in report.match_results
                    if r.status == "SUSPECT_MATCH")

    summary_data.extend([
        ("Total Matched Pairs", n_matched),
        ("NDF (Not in R2)", n_ndf),
        ("NEW (Only in R2)", n_new),
        ("", ""),
        ("LAYER 4 — GROWTH VALIDATION", ""),
        ("Active Corrosion", n_active),
        ("Stable", n_stable),
        ("Suspect Match", n_suspect),
    ])

    # Tulis summary
    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws_sum.cell(row=row_idx, column=1, value=label)
        ws_sum.cell(row=row_idx, column=2, value=value)
        # Style header rows
        if value == "" and label:
            ws_sum.cell(row=row_idx, column=1).font = Font(bold=True, size=12)

    # Format kolom
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 20

    # ERF critical list
    critical = [r for r in report.match_results
                if r.erf is not None and r.erf >= 1.0]
    if critical:
        row_start = len(summary_data) + 3
        ws_sum.cell(row=row_start, column=1,
                    value="⚠ CRITICAL ANOMALIES (ERF ≥ 1.0)")
        ws_sum.cell(row=row_start, column=1).font = Font(
            bold=True, color="FF0000", size=12)
        row_start += 1
        headers_crit = ["Anomaly ID (R2)", "Depth (%WT)", "ERF",
                        "Growth Rate (%/yr)", "Remaining Life (yr)"]
        for ci, h in enumerate(headers_crit, 1):
            c = ws_sum.cell(row=row_start, column=ci, value=h)
            c.fill = PatternFill(start_color="C00000", end_color="C00000",
                                 fill_type="solid")
            c.font = Font(bold=True, color="FFFFFF")
        for ri, mr in enumerate(critical, row_start + 1):
            ws_sum.cell(row=ri, column=1, value=mr.anom_id_r2)
            ws_sum.cell(row=ri, column=2,
                        value=round(mr.depth_r2, 1) if mr.depth_r2 else "")
            ws_sum.cell(row=ri, column=3,
                        value=round(mr.erf, 3) if mr.erf else "")
            ws_sum.cell(row=ri, column=4,
                        value=round(mr.growth_rate_per_yr, 3)
                        if mr.growth_rate_per_yr else "")
            ws_sum.cell(row=ri, column=5,
                        value=mr.remaining_life_yr if mr.remaining_life_yr else "")

    # ══════════════════════════════════════════════════════════════════════
    #  Sheet 2: Comparison (semua matched pairs)
    # ══════════════════════════════════════════════════════════════════════
    ws_comp = wb.create_sheet("Comparison")

    matched_results = [r for r in report.match_results
                       if r.status in ("MATCHED", "ACTIVE_CORROSION",
                                       "STABLE", "SUSPECT_MATCH")]
    if matched_results:
        df_comp = pd.DataFrame([{
            "spool_id":         mr.spool_id,
            "anom_id_R1":       mr.anom_id_r1,
            "anom_id_R2":       mr.anom_id_r2,
            "depth_R1 (%WT)":   mr.depth_r1,
            "depth_R2 (%WT)":   mr.depth_r2,
            "depth_growth (%)": mr.depth_growth_pct,
            "length_R1 (mm)":   mr.length_r1,
            "length_R2 (mm)":   mr.length_r2,
            "delta_odo (m)":    mr.delta_odo,
            "delta_clock (hr)": mr.delta_clock,
            "surface":          mr.side,
            "cost":             mr.cost,
            "status":           mr.status,
            "flag":             mr.flag,
            "growth_rate (%/yr)":  mr.growth_rate_per_yr,
            "ERF":              mr.erf,
            "remaining_life (yr)": mr.remaining_life_yr,
        } for mr in matched_results])
        _write_df_to_sheet(ws_comp, df_comp)
    else:
        ws_comp.cell(row=1, column=1, value="No matched pairs found")

    print(f"  Sheet Comparison: {len(matched_results)} rows")

    # ══════════════════════════════════════════════════════════════════════
    #  Sheet 3: Only_NEW (anomali hanya di R2)
    # ══════════════════════════════════════════════════════════════════════
    ws_new = wb.create_sheet("Only_NEW")

    new_results = [r for r in report.match_results if r.status == "NEW_IN_R2"]
    if new_results:
        df_new = pd.DataFrame([{
            "spool_id":       mr.spool_id,
            "anom_id_R2":     mr.anom_id_r2,
            "depth_R2 (%WT)": mr.depth_r2,
            "length_R2 (mm)": mr.length_r2,
            "surface":        mr.side,
            "status":         mr.status,
            "flag":           mr.flag,
        } for mr in new_results])
        _write_df_to_sheet(ws_new, df_new)
    else:
        ws_new.cell(row=1, column=1, value="No new anomalies found in R2")

    print(f"  Sheet Only_NEW: {len(new_results)} rows")

    # ══════════════════════════════════════════════════════════════════════
    #  Sheet 4: Only_OLD (anomali hanya di R1 — NDF)
    # ══════════════════════════════════════════════════════════════════════
    ws_old = wb.create_sheet("Only_OLD")

    ndf_results = [r for r in report.match_results if r.status == "NDF"]
    if ndf_results:
        df_ndf = pd.DataFrame([{
            "spool_id":       mr.spool_id,
            "anom_id_R1":     mr.anom_id_r1,
            "depth_R1 (%WT)": mr.depth_r1,
            "length_R1 (mm)": mr.length_r1,
            "surface":        mr.side,
            "status":         mr.status,
            "flag":           mr.flag,
        } for mr in ndf_results])
        _write_df_to_sheet(ws_old, df_ndf)
    else:
        ws_old.cell(row=1, column=1, value="No NDF anomalies found")

    print(f"  Sheet Only_OLD: {len(ndf_results)} rows")

    # ══════════════════════════════════════════════════════════════════════
    #  Sheet 5: Physics_Validation (Layer 4 results)
    # ══════════════════════════════════════════════════════════════════════
    ws_phys = wb.create_sheet("Physics_Validation")

    validated = [r for r in report.match_results
                 if r.status in ("ACTIVE_CORROSION", "STABLE", "SUSPECT_MATCH")]
    if validated:
        df_phys = pd.DataFrame([{
            "spool_id":         mr.spool_id,
            "anom_id_R1":       mr.anom_id_r1,
            "anom_id_R2":       mr.anom_id_r2,
            "depth_R1 (%WT)":   mr.depth_r1,
            "depth_R2 (%WT)":   mr.depth_r2,
            "depth_change (%)": mr.depth_growth_pct,
            "growth_rate (%/yr)":  mr.growth_rate_per_yr,
            "length_R1 (mm)":   mr.length_r1,
            "length_R2 (mm)":   mr.length_r2,
            "delta_odo (m)":    mr.delta_odo,
            "delta_clock (hr)": mr.delta_clock,
            "surface":          mr.side,
            "status":           mr.status,
            "flag":             mr.flag,
            "ERF":              mr.erf,
            "remaining_life (yr)": mr.remaining_life_yr,
        } for mr in validated])
        _write_df_to_sheet(ws_phys, df_phys)
    else:
        ws_phys.cell(row=1, column=1,
                     value="No physics-validated results available")

    print(f"  Sheet Physics_Validation: {len(validated)} rows")

    # ══════════════════════════════════════════════════════════════════════
    #  Sheet 6: Weld_Matching (Layer 2 details)
    # ══════════════════════════════════════════════════════════════════════
    ws_weld = wb.create_sheet("Weld_Matching")

    if report.weld_matching_summary:
        df_weld = pd.DataFrame(report.weld_matching_summary)
        _write_df_to_sheet(ws_weld, df_weld, apply_status_fill=False)
    else:
        ws_weld.cell(row=1, column=1, value="No weld matching data available")

    print(f"  Sheet Weld_Matching: {len(report.weld_matching_summary)} rows")

    # ══════════════════════════════════════════════════════════════════════
    #  Sheet 7: Valve_Correction (Layer 1 factors)
    # ══════════════════════════════════════════════════════════════════════
    ws_valve = wb.create_sheet("Valve_Correction")

    if report.correction_factors:
        df_valve = pd.DataFrame(report.correction_factors)
        _write_df_to_sheet(ws_valve, df_valve, apply_status_fill=False)
    else:
        ws_valve.cell(row=1, column=1,
                      value="No valve correction data available")

    print(f"  Sheet Valve_Correction: {len(report.correction_factors)} rows")

    # ── Simpan ─────────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"\n  ✓ Report saved: {output_path}")
    print(f"    File size: {os.path.getsize(output_path) / 1024:.1f} KB")


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 13 — MAIN PIPELINE ORCHESTRATION
# ══════════════════════════════════════════════════════════════════════════════════

def run_alignment_pipeline(
    file_r1: str,
    file_r2: str,
    output_path: str,
    year_r1: Optional[int] = None,
    year_r2: Optional[int] = None,
    vendor_r1: Optional[str] = None,
    vendor_r2: Optional[str] = None,
    **kwargs,
) -> AlignmentReport:
    """
    Orkestrasi utama: jalankan seluruh pipeline alignment secara berurutan.

    Sequence:
      1. Load kedua file Excel → Universal DataFrame
      2. Konversi ke run dictionary
      3. Layer 0: Validasi data
      4. Layer 1: Valve-to-valve odometer correction
      5. Layer 2: Weld-to-weld sequential matching
      6. Layer 3: Hungarian anomaly matching per spool
      7. Layer 4: Growth validation
      8. Tulis laporan Excel
      9. Return AlignmentReport

    Args:
        file_r1:     Path file Excel survey pertama (lebih lama)
        file_r2:     Path file Excel survey kedua (lebih baru)
        output_path: Path output file Excel laporan
        year_r1:     Tahun survey R1 (opsional, auto-detect dari nama file)
        year_r2:     Tahun survey R2 (opsional, auto-detect dari nama file)
        vendor_r1:   Vendor R1 (opsional)
        vendor_r2:   Vendor R2 (opsional)
        **kwargs:    Parameter tambahan (wt_mm, od_mm, smys_mpa, maop_bar, dll)

    Return:
        AlignmentReport dengan semua hasil
    """
    print("\n" + "═" * 70)
    print("  ILI PIPELINE ALIGNMENT SYSTEM v10.0")
    print("  4-Layer Sequential Hierarchical Engine")
    print("═" * 70)
    print(f"  R1: {os.path.basename(file_r1)}")
    print(f"  R2: {os.path.basename(file_r2)}")
    print(f"  Output: {output_path}")
    start_time = datetime.now()

    # ── Ambil parameter pipeline dari kwargs ───────────────────────────────
    wt_mm    = kwargs.get("wt_mm",    DEFAULT_WT_MM)
    od_mm    = kwargs.get("od_mm",    DEFAULT_OD_MM)
    smys_mpa = kwargs.get("smys_mpa", DEFAULT_SMYS_MPA)
    maop_bar = kwargs.get("maop_bar", DEFAULT_MAOP_BAR)
    max_weld_dist_m = kwargs.get("max_weld_dist_m", 3.0)

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 1: Load Excel files
    # ══════════════════════════════════════════════════════════════════════
    df_r1 = load_ili_universal(file_r1, survey_year=year_r1, vendor=vendor_r1)
    df_r2 = load_ili_universal(file_r2, survey_year=year_r2, vendor=vendor_r2)

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 2: Convert ke run dict
    # ══════════════════════════════════════════════════════════════════════
    run1 = df_to_run_dict(df_r1)
    run2 = df_to_run_dict(df_r2)

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 3: Layer 0 — Validasi data
    # ══════════════════════════════════════════════════════════════════════
    errors_r1 = layer0_validate(run1, label="R1")
    errors_r2 = layer0_validate(run2, label="R2")
    all_errors = errors_r1 + errors_r2

    if all_errors:
        print(f"\n  ⚠ {len(all_errors)} error validasi ditemukan, "
              f"melanjutkan dengan warning")

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 4: Layer 1 — Valve correction
    # ══════════════════════════════════════════════════════════════════════
    run1_c, run2_c, correction_factors = layer1_valve_correction(run1, run2)

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 5: Layer 2 — Weld matching
    # ══════════════════════════════════════════════════════════════════════
    spool_pairs, weld_summary, unmatched_w_r1, unmatched_w_r2 = \
        layer2_weld_matching(run1_c, run2_c, max_weld_dist_m=max_weld_dist_m)

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 6: Layer 3 — Anomaly matching
    # ══════════════════════════════════════════════════════════════════════
    match_results = layer3_anomaly_matching(run1_c, run2_c, spool_pairs)

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 7: Layer 4 — Growth validation
    # ══════════════════════════════════════════════════════════════════════
    # Tentukan tahun antar survey
    yr1 = year_r1 or df_r1["_survey_year"].iloc[0] if len(df_r1) > 0 else None
    yr2 = year_r2 or df_r2["_survey_year"].iloc[0] if len(df_r2) > 0 else None

    if yr1 and yr2:
        years_between = float(abs(yr2 - yr1))
    else:
        years_between = 5.0  # Default jika tahun tidak diketahui
        print(f"\n  ⚠ Tahun survey tidak terdeteksi, "
              f"menggunakan default {years_between} tahun")

    match_results = layer4_growth_validation(
        match_results, years_between,
        wt_mm=wt_mm, od_mm=od_mm, smys_mpa=smys_mpa, maop_bar=maop_bar,
    )

    # ══════════════════════════════════════════════════════════════════════
    #  Build AlignmentReport
    # ══════════════════════════════════════════════════════════════════════
    report = AlignmentReport(
        n_valves_r1=len(run1["valves"]),
        n_valves_r2=len(run2["valves"]),
        n_welds_matched=len(weld_summary),
        n_welds_unmatched_r1=len(unmatched_w_r1),
        n_welds_unmatched_r2=len(unmatched_w_r2),
        n_spool_pairs=len(spool_pairs),
        match_results=match_results,
        correction_factors=correction_factors,
        weld_matching_summary=weld_summary,
        errors=all_errors,
        _corrected_run1=run1_c,
        _corrected_run2=run2_c,
        _spool_pairs=spool_pairs,
    )

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 8: Write Excel report
    # ══════════════════════════════════════════════════════════════════════
    write_excel_report(report, output_path, df_r1=df_r1, df_r2=df_r2)

    # ══════════════════════════════════════════════════════════════════════
    #  STEP 9: Summary & Return
    # ══════════════════════════════════════════════════════════════════════
    elapsed = (datetime.now() - start_time).total_seconds()
    print(f"\n{'═'*70}")
    print(f"  PIPELINE COMPLETE")
    print(f"  Elapsed: {elapsed:.1f}s")
    print(f"  Results: {len(match_results)} total entries")
    print(f"  Report:  {output_path}")
    print(f"{'═'*70}\n")

    return report


# ══════════════════════════════════════════════════════════════════════════════════
#  BAGIAN 14 — HELPER: results_to_dataframe
# ══════════════════════════════════════════════════════════════════════════════════

def results_to_dataframe(results: list[MatchResult]) -> pd.DataFrame:
    """
    Konversi list[MatchResult] ke pandas DataFrame.

    Semua field dari MatchResult disertakan sebagai kolom.
    Berguna untuk analisis lanjutan, filtering, dan visualisasi.
    """
    if not results:
        return pd.DataFrame(columns=[
            "spool_id", "anom_id_r1", "anom_id_r2",
            "delta_odo", "delta_clock",
            "depth_r1", "depth_r2",
            "length_r1", "length_r2",
            "side", "cost", "status", "flag",
            "depth_growth_pct", "growth_rate_per_yr",
            "erf", "remaining_life_yr",
        ])

    records = []
    for mr in results:
        records.append({
            "spool_id":          mr.spool_id,
            "anom_id_r1":        mr.anom_id_r1,
            "anom_id_r2":        mr.anom_id_r2,
            "delta_odo":         mr.delta_odo,
            "delta_clock":       mr.delta_clock,
            "depth_r1":          mr.depth_r1,
            "depth_r2":          mr.depth_r2,
            "length_r1":         mr.length_r1,
            "length_r2":         mr.length_r2,
            "side":              mr.side,
            "cost":              mr.cost,
            "status":            mr.status,
            "flag":              mr.flag,
            "depth_growth_pct":  mr.depth_growth_pct,
            "growth_rate_per_yr": mr.growth_rate_per_yr,
            "erf":               mr.erf,
            "remaining_life_yr": mr.remaining_life_yr,
        })

    df = pd.DataFrame(records)

    # ── Urutkan: ACTIVE_CORROSION pertama, lalu SUSPECT, STABLE, NDF, NEW ──
    status_order = {
        "ACTIVE_CORROSION": 0,
        "SUSPECT_MATCH":    1,
        "STABLE":           2,
        "MATCHED":          3,
        "NDF":              4,
        "NEW_IN_R2":        5,
        "UNVALIDATED":      6,
    }
    df["_sort_key"] = df["status"].map(status_order).fillna(99)
    df = df.sort_values(["_sort_key", "spool_id"]).drop(columns=["_sort_key"])
    df = df.reset_index(drop=True)

    return df


# ══════════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT (CLI)
# ══════════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    """
    Penggunaan CLI:
        python ili_engine.py <file_r1.xlsx> <file_r2.xlsx> [output.xlsx]
                             [--year-r1 YYYY] [--year-r2 YYYY]
                             [--wt WT_MM] [--od OD_MM]
                             [--smys SMYS_MPA] [--maop MAOP_BAR]
    """
    import argparse

    parser = argparse.ArgumentParser(
        description="ILI Pipeline Alignment System v10.0"
    )
    parser.add_argument("file_r1", help="Path file Excel survey R1 (lama)")
    parser.add_argument("file_r2", help="Path file Excel survey R2 (baru)")
    parser.add_argument("output", nargs="?", default=None,
                        help="Path output Excel (default: alignment_report.xlsx)")
    parser.add_argument("--year-r1", type=int, default=None,
                        help="Tahun survey R1")
    parser.add_argument("--year-r2", type=int, default=None,
                        help="Tahun survey R2")
    parser.add_argument("--vendor-r1", type=str, default=None,
                        help="Vendor R1")
    parser.add_argument("--vendor-r2", type=str, default=None,
                        help="Vendor R2")
    parser.add_argument("--wt", type=float, default=DEFAULT_WT_MM,
                        help=f"Wall thickness (mm), default={DEFAULT_WT_MM}")
    parser.add_argument("--od", type=float, default=DEFAULT_OD_MM,
                        help=f"Outside diameter (mm), default={DEFAULT_OD_MM}")
    parser.add_argument("--smys", type=float, default=DEFAULT_SMYS_MPA,
                        help=f"SMYS (MPa), default={DEFAULT_SMYS_MPA}")
    parser.add_argument("--maop", type=float, default=DEFAULT_MAOP_BAR,
                        help=f"MAOP (bar), default={DEFAULT_MAOP_BAR}")
    parser.add_argument("--max-weld-dist", type=float, default=3.0,
                        help="Max weld matching distance (m), default=3.0")

    args = parser.parse_args()

    # Default output path
    if args.output is None:
        out_dir = os.path.dirname(os.path.abspath(args.file_r1))
        args.output = os.path.join(out_dir, "alignment_report.xlsx")

    report = run_alignment_pipeline(
        file_r1=args.file_r1,
        file_r2=args.file_r2,
        output_path=args.output,
        year_r1=args.year_r1,
        year_r2=args.year_r2,
        vendor_r1=args.vendor_r1,
        vendor_r2=args.vendor_r2,
        wt_mm=args.wt,
        od_mm=args.od,
        smys_mpa=args.smys,
        maop_bar=args.maop,
        max_weld_dist_m=args.max_weld_dist,
    )

    # Print ringkasan akhir
    df_results = results_to_dataframe(report.match_results)
    print("\nRingkasan Hasil:")
    print(df_results["status"].value_counts().to_string())
